import pandas as pd
from openpyxl import Workbook
#from phi.flow import *
from phi.jax.flow import *
import matplotlib

def main():
    
    d=2 # dimension of the problem (indirectly dimension of the kernal function) 
    dx=0.006 #distance between particles
    h=dx #cut off radius
    r_c=3*h   #Quintic Spline
    alpha=0.02 #viscosity coefficient value of water

    fluid_particles,wall_particles, fluid_initial_density,wall_initial_density, \
    fluid_particle_density,wall_particle_density, fluid_particle_pressure, wall_particle_pressure,  \
    fluid_particle_mass, \
    fluid_adiabatic_exp,wall_adiabatic_exp, fluid_c_0,wall_c_0,fluid_p_0,wall_p_0,fluid_Xi,wall_Xi,fluid_alpha,wall_alpha, \
    g, Height= intial_condition(dx,d,alpha)

    t=0
    n_dt=0

    ################################################
    ### To be corrected
    ################################################
    H=math.max(fluid_particles.points['y'])-math.min(fluid_particles.points['y'])+dx
    math.print(H.all)
    print(Height)
    print('Actual height of water column is: ' +str(H))
    if abs(H.all - Height) >= 1.0e-6:
        print("wrong height specified")
        print(H.any-Height)
        #exit()
    
    #reference_value
    v_ref = math.sqrt(2 *abs(g) * H )
    t_ref = H / v_ref 

    time_nondim = 5    # nondimensional running time (t/t_ref)
    t_end = numpy.asarray(time_nondim * t_ref)

    print("Total simulation time is: "+str(t_end))
    print("Total number of fluid particles: "+ str(fluid_particles.elements.center.particles.size))
    print("Total number of wall particles: "+ str(wall_particles.elements.center.particles.size))
    
    fluid_traj=[]
    while n_dt <=11600:

        print("Simulation progress: "+ str((100*(t/t_end)))+ " Time step "+ str(n_dt))
        dt = time_step_size(fluid_c_0,fluid_particles, wall_particles,h,fluid_alpha,d,g)
        print("Time step size is: "+str(dt))
        
        t=t+dt
        n_dt=n_dt+1

        if n_dt == 1:

            wall_particle_pressure, wall_particles =boundary_update(fluid_particles,wall_particles,fluid_particle_pressure,fluid_particle_density,d,r_c,h,g)
            fluid_particle_acceleration =calculate_acceleration(fluid_particles, wall_particles, fluid_particle_density,fluid_particle_pressure,wall_particle_pressure,fluid_particle_mass,fluid_initial_density,fluid_Xi,fluid_adiabatic_exp,fluid_p_0, h, d, r_c, dx, fluid_alpha,fluid_c_0,g,n_dt)

            workbook_y = Workbook()
            workbook_y.save("y_pos.xlsx")

            workbook_x = Workbook()
            workbook_x.save("x_pos.xlsx")

            pos_x = numpy.asarray(math.concat([fluid_particles.points['x'],wall_particles.points['x']],'particles'))
            pos_y = numpy.asarray(math.concat([fluid_particles.points['y'],wall_particles.points['y']],'particles'))
            df_x=pd.DataFrame(pos_x)
            df_y=pd.DataFrame(pos_y)

            with pd.ExcelWriter("x_pos.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer_x:
                df_x.to_excel(writer_x, sheet_name="Sheet",header=None, startcol=writer_x.sheets["Sheet"].max_column,index=False)
                writer_x.save()

            with pd.ExcelWriter("y_pos.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer_y:
                df_y.to_excel(writer_y, sheet_name="Sheet",header=None, startcol=writer_y.sheets["Sheet"].max_column,index=False)
                writer_y.save()

        #*******************************************************************************************
        # Step-Function: Calculate the parameters for one step
        #*******************************************************************************************
        fluid_particles,wall_particles, fluid_particle_acceleration, fluid_particle_pressure, fluid_particle_density, wall_particle_pressure = step(fluid_particles,wall_particles,fluid_particle_acceleration,fluid_particle_pressure,wall_particle_pressure,fluid_initial_density,fluid_particle_density,wall_particle_density, fluid_particle_mass, \
        fluid_adiabatic_exp, fluid_c_0,fluid_p_0,fluid_Xi,fluid_alpha, dt,n_dt, d,r_c,h,g,dx)
        #*******************************************************************************************
          
        #fluid_traj.append(fluid_particles)

        if(n_dt%150==0):
            pos_x = numpy.asarray(math.concat([fluid_particles.points['x'],wall_particles.points['x']],'particles'))
            pos_y = numpy.asarray(math.concat([fluid_particles.points['y'],wall_particles.points['y']],'particles'))
            df_x=pd.DataFrame(pos_x)
            df_y=pd.DataFrame(pos_y)

            with pd.ExcelWriter("x_pos.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer_x:
                df_x.to_excel(writer_x, sheet_name="Sheet",header=None, startcol=writer_x.sheets["Sheet"].max_column,index=False)
                writer_x.save()

            with pd.ExcelWriter("y_pos.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer_y:
                df_y.to_excel(writer_y, sheet_name="Sheet",header=None, startcol=writer_y.sheets["Sheet"].max_column,index=False)
                writer_y.save()
        
    #**************************
    # Save animation
    #**************************
    # fluid_trj = math.stack(fluid_traj, batch('time'))
    # a: matplotlib.animation.FuncAnimation = vis.plot(vis.overlay(wall_particles.elements, fluid_trj.elements), animate='time')
    # a.save('anim.mp4')

def step(fluid_particles,wall_particles,fluid_particle_acceleration,fluid_particle_pressure,wall_particle_pressure,fluid_initial_density,fluid_particle_density,wall_particle_density, fluid_particle_mass, \
    fluid_adiabatic_exp, fluid_c_0,fluid_p_0,fluid_Xi,fluid_alpha, dt,n_dt,d,r_c,h,g,dx):

    fluid_particle_velocity=math.expand(fluid_particles.values, instance(fluid_particles))

    fluid_particle_velocity = fluid_particle_velocity + (dt/2)*fluid_particle_acceleration #initial fluid_particle_velocity obtained from dam_break_case function
    fluid_particle_position = fluid_particles.points + (dt/2)*fluid_particle_velocity
    #update the point cloud. positions are stored in elements of the pointcloudand velocities are stored in values of the pointcloud
    fluid_particles = fluid_particles.with_values(fluid_particle_velocity)
    fluid_particles = fluid_particles.with_elements(fluid_particle_position)
    
    wall_particle_pressure, wall_particles = boundary_update(fluid_particles,wall_particles,fluid_particle_pressure,fluid_particle_density,d,r_c,h,g)
    
    drho_by_dt= calculate_density_derivative(fluid_particles, wall_particles, fluid_particle_density,fluid_particle_pressure,fluid_particle_mass,fluid_initial_density,fluid_Xi,fluid_adiabatic_exp,fluid_p_0, h, d, r_c, dx)
    
    #Density Update: 
    fluid_particle_density = fluid_particle_density + (dt*drho_by_dt)

    #Pressure Update:  p_0[0] * ((particles[:len(fluid_particle_indices),5]/rho_0[0])**gamma[0] - 1 ) + Xi[0] 
    fluid_particle_pressure = fluid_p_0 *((fluid_particle_density/fluid_initial_density)**fluid_adiabatic_exp -1) + fluid_Xi

    fluid_particle_position = fluid_particles.points + (dt/2)*fluid_particle_velocity
    fluid_particles = fluid_particles.with_elements(fluid_particle_position)

    wall_particle_pressure, wall_particles= boundary_update(fluid_particles,wall_particles,fluid_particle_pressure,fluid_particle_density,d,r_c,h,g)

    #Calculate Acceleration
    fluid_particle_acceleration = calculate_acceleration(fluid_particles, wall_particles, fluid_particle_density,fluid_particle_pressure,wall_particle_pressure,fluid_particle_mass,fluid_initial_density,fluid_Xi,fluid_adiabatic_exp,fluid_p_0, h, d, r_c, dx, fluid_alpha,fluid_c_0,g,n_dt)
    
    fluid_particle_velocity = fluid_particle_velocity + (dt/2)*fluid_particle_acceleration
    fluid_particles = fluid_particles.with_values(fluid_particle_velocity)

    return fluid_particles,wall_particles, fluid_particle_acceleration, fluid_particle_pressure, fluid_particle_density, wall_particle_pressure

step = jit_compile(step, auxiliary_args='d')

def boundary_update(fluid_particles,wall_particles,fluid_pressure,fluid_density,d,r_c,h,g):

    fluid_coords=fluid_particles.points   #fluid_particles and wall_particles are a point cloud objects
    wall_coords=wall_particles.points

    particle_coords=math.concat([fluid_coords,wall_coords],'particles')  # concatenating fluid coords and then wall coords

    distance_matrix_vec= particle_coords - math.rename_dims(particle_coords, 'particles', 'others') # contains both the x and y component of separation between particles
    distance_matrix = math.vec_length(distance_matrix_vec) # contains magnitude of distance between particles

    #Apply cut off to the distance matrix at the end after all the cut-offs
    distance_matrix_vec= math.where(distance_matrix > r_c, 0, distance_matrix_vec)
    distance_matrix = math.where(distance_matrix > r_c, 0, distance_matrix)

    #Slicing the distance matrix of fluid neighbour of wall particles
    q=distance_matrix.particles[fluid_coords.particles.size:].others[:fluid_coords.particles.size]/h   

    W=kernal_function(d,h,q)

    sum_pW = W.others*fluid_pressure.particles  # col-vector, each entry corresponding to each wall particle

    dry=distance_matrix_vec['y'].particles[fluid_coords.particles.size:].others[:fluid_coords.particles.size]*g 
    drx=distance_matrix_vec['x'].particles[fluid_coords.particles.size:].others[:fluid_coords.particles.size]*0 

    sum_rhorW =(dry*W).others*fluid_density.particles # col-vector, each entry corresponding to each wall particle

    sum_W = math.sum(W,'others')  # col-vector, (row sum of W) each entry corresponding to each wall particle

    wall_particle_pressure= math.where(sum_W==0,0,sum_W)
    wall_particle_pressure= math.where(sum_W!=0,math.divide_no_nan((sum_rhorW+sum_pW),sum_W),sum_W)

    fluid_particle_velocity=math.expand(fluid_particles.values, instance(fluid_particles))
    wall_particle_velocity=math.expand(wall_particles.values, instance(wall_particles))

    sum_vWx = W.others*fluid_particle_velocity['x'].particles
    sum_vWy = W.others*fluid_particle_velocity['y'].particles
    
    #sum_W and wall_particle_velocity['x' or 'y'] have the same dimensions i.e. both are a column vector with particle dimension
    wall_vel_x = math.where(sum_W!=0, math.divide_no_nan(-sum_vWx,sum_W), wall_particle_velocity['x']) 
    wall_vel_y = math.where(sum_W!=0, math.divide_no_nan(-sum_vWy,sum_W), wall_particle_velocity['y'])
    
    wall_particle_velocity_temp = stack([wall_vel_x, wall_vel_y], channel(vector='x,y'))
        
    wall_particles = wall_particles.with_values(wall_particle_velocity_temp) # wall_particle_velocity is a point cloud
    
    return wall_particle_pressure, wall_particles

def calculate_density_derivative(fluid_particles, wall_particles, fluid_particle_density,fluid_pressure,fluid_particle_mass,fluid_initial_density,fluid_Xi,fluid_adiabatic_exp,fluid_p_0, h, d, r_c, dx):
    #Note: No wall_particle_density and wall_particle_mass in input arguments. It is expected to be zero initially
    fluid_coords=fluid_particles.points   #fluid_particles and boundary_particles are a point cloud objects
    wall_coords=wall_particles.points
    particle_coords=math.concat([fluid_coords,wall_coords],'particles')  # concatenating fluid coords and then boundary coords
    
    fluid_particle_velocity=math.expand(fluid_particles.values, instance(fluid_particles)) # adding particle dimension to fluid_particles.values
    wall_particle_velocity = math.expand(wall_particles.values, instance(wall_particles))

    particle_velocity=math.concat([fluid_particle_velocity,wall_particle_velocity], dim='particles') 

    wall_particle_density=math.rename_dims(math.zeros(instance(wall_coords)),'particles','others')
    wall_particle_mass=math.rename_dims(math.zeros(instance(wall_coords)),'particles','others')

    fluid_particle_density=math.rename_dims(fluid_particle_density,'particles','others') # fluid_particle_density and mass have initially particle dimension (i.e column vector)
    fluid_particle_mass=math.rename_dims(fluid_particle_mass,'particles','others')
    
    particle_density= math.concat([fluid_particle_density,wall_particle_density], dim='others') #1D Scalar array of densities of all particles (now it is a row vector)
    particle_mass= math.concat([fluid_particle_mass,wall_particle_mass], dim='others')
    
    #Compute distance between all particles 
    distance_matrix_vec= particle_coords - math.rename_dims(particle_coords, 'particles', 'others') # contains both the x and y component of separation between particles
    distance_matrix = math.vec_length(distance_matrix_vec) # contains magnitude of distance between ALL particles
    
    particle_neighbour_density = math.where(distance_matrix==0,0,particle_density)
    particle_neighbour_density=math.where(distance_matrix > r_c, 0, particle_neighbour_density) #0 for places which are not neighbours for particle under consideration. Stacks the particle density vertically. i.e. dupicates the densities
    
    particle_neighbour_mass = math.where(distance_matrix==0,0,particle_mass)
    particle_neighbour_mass=math.where(distance_matrix > r_c, 0, particle_neighbour_mass)

    particle_relative_velocity=particle_velocity - math.rename_dims(particle_velocity,'particles', 'others')# 2d matrix of ALL particle velocity
    dvx=particle_relative_velocity['x'].particles[:fluid_coords.particles.size].others[:]# separating the x and y components
    dvy=particle_relative_velocity['y'].particles[:fluid_coords.particles.size].others[:]

    fluid_particle_relative_dist=distance_matrix.particles[:fluid_coords.particles.size].others[:]
    dvx=math.where(fluid_particle_relative_dist>r_c,0,dvx) # relative x-velocity between a fluid particle and its neighbour
    dvy=math.where(fluid_particle_relative_dist>r_c,0,dvy)
 
    #Do all the cut off things before the final cut off for distance matrix
    distance_matrix_vec= math.where(distance_matrix > r_c, 0, distance_matrix_vec)
    distance_matrix = math.where(distance_matrix > r_c, 0, distance_matrix)   # Stores the distance between neighbours which are inside cutoff radius

    #Slicing the distance matrix of ALL neighbour of fluid particles
    q=distance_matrix.particles[:fluid_coords.particles.size].others[:]/h 

    DWab = kernal_derivative(d, h, q) / h

    drx=distance_matrix_vec['x'].particles[:fluid_coords.particles.size].others[:]
    dry=distance_matrix_vec['y'].particles[:fluid_coords.particles.size].others[:] 

    mod_dist=(distance_matrix.particles[:fluid_coords.particles.size].others[:])
    Fab_x= math.where(mod_dist!=0, math.divide_no_nan(drx,mod_dist)*DWab,drx) 
    Fab_y= math.where(mod_dist!=0, math.divide_no_nan(dry,mod_dist)*DWab,dry)

    fluid_particle_neighbour_density=particle_neighbour_density.particles[:fluid_coords.particles.size].others[:]
    fluid_particle_neighbour_mass=particle_neighbour_mass.particles[:fluid_coords.particles.size].others[:]

    #Splitting the density and mass of fluid particle neighbours into two matrices....QQQQQQ CAN WE AVOID THIS? as we have to join it back anyways
    fluid_particle_fluid_neighbour_density=fluid_particle_neighbour_density.particles[:].others[:fluid_coords.particles.size]
    fluid_particle_wall_neighbour_density=fluid_particle_neighbour_density.particles[:].others[fluid_coords.particles.size:]
    fluid_particle_fluid_neighbour_mass=fluid_particle_neighbour_mass.particles[:].others[:fluid_coords.particles.size]
    fluid_particle_wall_neighbour_mass=fluid_particle_neighbour_mass.particles[:].others[fluid_coords.particles.size:]

    ######'fluid_pressure' is for each fluid particle 
    fluid_particle_wall_neighbour_density_term=fluid_initial_density * ((fluid_pressure - fluid_Xi) / fluid_p_0 + 1)**(1/fluid_adiabatic_exp)
    helper_matrix = distance_matrix.particles[:fluid_coords.particles.size].others[fluid_coords.particles.size:] 
    helper_matrix = math.where(helper_matrix!=0, 1 , helper_matrix)  # technically called adjacency matrix
    
    #ALL boundary particle neighbours for a given fluid particle will have same density 
    fluid_particle_wall_neighbour_density = helper_matrix * fluid_particle_wall_neighbour_density_term
    fluid_particle_wall_neighbour_mass =math.where(helper_matrix!=0,(fluid_initial_density*dx*dx),fluid_particle_wall_neighbour_mass)
    
    #joining the density and mass of the fluid and wall particle back together
    fluid_particle_neighbour_density=math.concat([fluid_particle_fluid_neighbour_density, fluid_particle_wall_neighbour_density], 'others')
    fluid_particle_neighbour_mass =math.concat([fluid_particle_fluid_neighbour_mass, fluid_particle_wall_neighbour_mass], 'others')

    neighbour_mass_by_density_ratio=math.divide_no_nan(fluid_particle_neighbour_mass,fluid_particle_neighbour_density) #m_b/rho_b
    dot_product_result=(dvx*Fab_x)+(dvy*Fab_y)

    #Fluid_particle_density is a vector, each row of result is multiplied by the corresponding term of fluid_particle_density and then Sum along each row to get the result(drho_by_dt) for each fluid particle 
    fluid_particle_density= math.rename_dims(fluid_particle_density,'others','particles') 

    drho_by_dt=math.sum((neighbour_mass_by_density_ratio*dot_product_result)*fluid_particle_density,'others')
    #drho_by_dt is calculated only for fluid particles
    return drho_by_dt

def calculate_acceleration(fluid_particles, wall_particles ,fluid_particle_density,fluid_particle_pressure,wall_particle_pressure,fluid_particle_mass,fluid_initial_density,fluid_Xi,fluid_adiabatic_exp,fluid_p_0, h, d, r_c, dx, fluid_alpha,fluid_c_0,g,n_dt):

    dx=h; epsilon= 0.01 # parameter to avoid zero denominator 

    fluid_coords=fluid_particles.points   #fluid_particles and boundary_particles are a point cloud objects
    wall_coords=wall_particles.points
    particle_coords=math.concat([fluid_coords,wall_coords],'particles')  # concatenating fluid coords and then boundary coords
    
    fluid_particle_density = math.rename_dims(fluid_particle_density,'particles','others')
    wall_particle_density=math.rename_dims(math.zeros(instance(wall_coords)),'particles','others')
    
    fluid_particle_mass = math.rename_dims(fluid_particle_mass,'particles', 'others')
    wall_particle_mass=math.rename_dims(math.zeros(instance(wall_coords)),'particles','others')
    
    fluid_particle_velocity=math.expand(fluid_particles.values, instance(fluid_particles))
    wall_particle_velocity=math.expand(wall_particles.values, instance(wall_particles))

    fluid_particle_velocity = math.rename_dims(fluid_particle_velocity, 'particles', 'others')
    wall_particle_velocity=math.rename_dims(wall_particle_velocity,'particles','others')

    fluid_particle_pressure = math.rename_dims(fluid_particle_pressure, 'particles', 'others')
    wall_particle_pressure=math.rename_dims(wall_particle_pressure,'particles','others')
    
    particle_velocity = math.concat([fluid_particle_velocity,wall_particle_velocity], dim='others')
    particle_density= math.concat([fluid_particle_density,wall_particle_density], dim='others') #1D Scalar array of densities of all particles
    particle_mass= math.concat([fluid_particle_mass,wall_particle_mass], dim='others')
    particle_pressure=math.concat([fluid_particle_pressure,wall_particle_pressure], dim='others')

    #Compute distance between all particles 
    distance_matrix_vec= particle_coords - math.rename_dims(particle_coords, 'particles', 'others') # contains both the x and y component of separation between particles
    distance_matrix = math.vec_length(distance_matrix_vec) # contains magnitude of distance between ALL particles

    alpha_ab = math.where(distance_matrix==0,0,fluid_alpha) #Removing the particle itself from further calculation
    alpha_ab=math.where(distance_matrix> r_c, 0, alpha_ab)   #N X N matrix N-->all particles
    fluid_particle_neighbour_alpha_ab=alpha_ab.particles[:fluid_coords.particles.size].others[:]
    
    c_ab = math.where(distance_matrix==0,0,fluid_c_0) #Removing the particle itself from further calculation
    c_ab=math.where(distance_matrix> r_c, 0, c_ab)   #N X N matrix N-->all particles
    fluid_particle_neighbour_c_ab=c_ab.particles[:fluid_coords.particles.size].others[:]

    #Below we create matrices which have non-zero entries where the neighbour is inside cut-off radius 'r_c' rest all entries are 0
    particle_velocity=math.rename_dims(particle_velocity,'others','particles')
    particle_velocity_matrix =  particle_velocity- math.rename_dims(particle_velocity, 'particles', 'others') 

    particle_neighbour_density = math.where(distance_matrix==0,0,particle_density)  #Removing the particle itself from further calculation
    particle_neighbour_density = math.where(distance_matrix > r_c, 0, particle_neighbour_density) #0 for places which are not neighbours for particle under consideration. Stacks the particle density vertically. i.e. dupicates the densities
    
    particle_neighbour_mass = math.where(distance_matrix==0,0,particle_mass) #Removing the particle itself from further calculation
    particle_neighbour_mass = math.where(distance_matrix > r_c, 0, particle_neighbour_mass)

    particle_neighbour_pressure = math.where(distance_matrix==0,0,particle_pressure) #Removing the particle itself from further calculation
    particle_neighbour_pressure= math.where(distance_matrix > r_c,0, particle_neighbour_pressure)

    dvx=particle_velocity_matrix['x'].particles[:fluid_coords.particles.size].others[:]# separating the x and y components
    dvy=particle_velocity_matrix['y'].particles[:fluid_coords.particles.size].others[:]

    fluid_particle_all_neighbours_dist=distance_matrix.particles[:fluid_coords.particles.size].others[:]
    dvx=math.where(fluid_particle_all_neighbours_dist>r_c,0,dvx) # relative x-velocity between a fluid particle and its neighbour
    dvy=math.where(fluid_particle_all_neighbours_dist>r_c,0,dvy)

    ######Do all the cut off things before the final cut off for distance matrix
    distance_matrix_vec= math.where(distance_matrix > r_c, 0, distance_matrix_vec)
    distance_matrix = math.where(distance_matrix > r_c, 0, distance_matrix)   # Stores the distance between neighbours which are inside cutoff radius

    #Slicing the distance matrix of fluid neighbour of fluid particles
    rad=distance_matrix.particles[:fluid_coords.particles.size].others[:]

    q=rad/h 
  
    der_W = kernal_derivative(d, h, q) / h

    drx=distance_matrix_vec['x'].particles[:fluid_coords.particles.size].others[:]
    dry=distance_matrix_vec['y'].particles[:fluid_coords.particles.size].others[:] 
    
    #rho_b, m_b and p_b matrices rows are for each fluid particle and cols are fluid and wall particles
    ########THE BELOW 3 MATRICES CAN BE REMOVED
    fluid_particle_neighbour_density=particle_neighbour_density.particles[:fluid_coords.particles.size].others[:]
    fluid_particle_neighbour_mass=particle_neighbour_mass.particles[:fluid_coords.particles.size].others[:]
    fluid_particle_neighbour_pressure = particle_neighbour_pressure.particles[:fluid_coords.particles.size].others[:]

    #Splitting the density and mass of fluid particle neighbours into two matrices....QQQQQQ CAN WE AVOID THIS? as we have to join it back anyways
    fluid_particle_fluid_neighbour_density=fluid_particle_neighbour_density.particles[:].others[:fluid_coords.particles.size]
    fluid_particle_wall_neighbour_density=fluid_particle_neighbour_density.particles[:].others[fluid_coords.particles.size:]
    fluid_particle_fluid_neighbour_mass=fluid_particle_neighbour_mass.particles[:].others[:fluid_coords.particles.size]
    fluid_particle_wall_neighbour_mass=fluid_particle_neighbour_mass.particles[:].others[fluid_coords.particles.size:]

    ######'fluid_particle_pressure' is calculated and stored for each fluid particle 
    fluid_particle_wall_neighbour_density_term=fluid_initial_density * ((fluid_particle_pressure - fluid_Xi) / fluid_p_0 + 1)**(1/fluid_adiabatic_exp)

    fluid_particle_wall_neighbour_density_term= math.rename_dims(fluid_particle_wall_neighbour_density_term,'others','particles')
    helper_matrix = distance_matrix.particles[:fluid_coords.particles.size].others[fluid_coords.particles.size:] 
    helper_matrix = math.where(helper_matrix!=0, 1 , helper_matrix)  # technically called adjacency matrix

    #ALL boundary particle neighbours for a given fluid particle will have same density 
    fluid_particle_wall_neighbour_density = helper_matrix * fluid_particle_wall_neighbour_density_term

    #fluid_particle_wall_neighbour_density=math.where(fluid_particle_wall_neighbour_density!=0,fluid_particle_wall_neighbour_density_term,fluid_particle_wall_neighbour_density)
    fluid_particle_wall_neighbour_mass =math.where(helper_matrix!=0,(fluid_initial_density*dx*dx),fluid_particle_wall_neighbour_mass)

    #joining the density and mass of the fluid and wall particle back together
    fluid_particle_neighbour_density=math.concat([fluid_particle_fluid_neighbour_density, fluid_particle_wall_neighbour_density], 'others')
    fluid_particle_neighbour_mass =math.concat([fluid_particle_fluid_neighbour_mass, fluid_particle_wall_neighbour_mass], 'others')

    #Momentum Equation for the pressure gradient part
    fluid_particle_density= math.rename_dims(fluid_particle_density,'others','particles') 
    fluid_particle_pressure = math.rename_dims(fluid_particle_pressure, 'others', 'particles')
    fluid_particle_mass = math.rename_dims(fluid_particle_mass,'others','particles')
        
    #(rho_a + rho_b) 
    particle_density_sum=math.where(fluid_particle_neighbour_density!=0, fluid_particle_density+fluid_particle_neighbour_density,fluid_particle_neighbour_density)

    #rho_ab = 0.5 * (rho_a + rho_b)
    rho_ab=0.5*particle_density_sum

    #Setting minimum density as the initial density
    fluid_particle_density=math.where(fluid_particle_density==0,fluid_initial_density,fluid_particle_density)

    #p_ab = ((rho_b * p_a) + (rho_a * p_b)) / (rho_a + rho_b) 
    term_1 = fluid_particle_neighbour_density*fluid_particle_pressure
    term_2 = fluid_particle_neighbour_pressure*fluid_particle_density
    p_ab= math.divide_no_nan(((term_1) + (term_2)),(particle_density_sum))

    #pressure_fact = - (1/m_a) * ((m_a/rho_a)**2 + (m_b/rho_b)**2) * (p_ab) * der_W #equation 5
    fluid_neighbour_mass_by_density_ratio=math.divide_no_nan(fluid_particle_neighbour_mass,fluid_particle_neighbour_density)**2 #(m_b/rho_b)² 
    fluid_mass_by_density_ratio=math.divide_no_nan(fluid_particle_mass,fluid_particle_density)**2  # (m_a/rho_a)²

    #sum = (m_a/rho_a)**2 + (m_b/rho_b)**2
    mass_by_density_sum =math.where(fluid_neighbour_mass_by_density_ratio!=0,fluid_mass_by_density_ratio+fluid_neighbour_mass_by_density_ratio,fluid_neighbour_mass_by_density_ratio)

    # pressure_fact=-(1/m_a) * ((m_a/rho_a)**2 + (m_b/rho_b)**2) * (p_ab) * der_W (equation 5)
    pressure_fact=(-1/fluid_particle_mass)*mass_by_density_sum*p_ab*der_W

    #a_x
    a_x=math.sum(pressure_fact*math.divide_no_nan(drx,rad),'others')
    #a_y
    a_y=math.sum(pressure_fact*math.divide_no_nan(dry,rad),'others')
   
    ### ARTIFICIAL VISCOSITY
    visc_art_fact=math.zeros(instance(fluid_particle_all_neighbours_dist))  #zeros matrix with size: (fluid_particles X all_particles)
    temp_matrix= (drx*dvx) + (dry*dvy) 
    
    # visc_art_fact_term = m_b * alpha_ab * h * c_ab * (((dvx * drx) + (dvy * dry))/(rho_ab * ((rad**2) + epsilon * (h**2)))) * der_W
    visc_art_fact_term = fluid_particle_neighbour_mass * fluid_particle_neighbour_alpha_ab * h * fluid_particle_neighbour_c_ab* math.divide_no_nan((temp_matrix),(rho_ab * ((rad**2) + epsilon * (h**2)))) * der_W
    visc_art_fact = math.where(temp_matrix<0, visc_art_fact_term, visc_art_fact)
    
    #a_x
    a_x=a_x+math.sum(visc_art_fact*math.divide_no_nan(drx,rad),'others')
    #a_y
    a_y=a_y+math.sum(visc_art_fact*math.divide_no_nan(dry,rad),'others')

    #### GRAVITY
    a_y=a_y+g

    fluid_particle_acceleration = stack([a_x, a_y], channel(vector='x,y'))
    #Acceleration of only fluid particles calculated and returned
    
    return fluid_particle_acceleration

def time_step_size(fluid_c_0,fluid_particles, wall_particles,h,fluid_alpha,d,g):

    fluid_particle_velocity=math.expand(fluid_particles.values, instance(fluid_particles)) # adding particle dimension to fluid_particles.values
    wall_particle_velocity = math.expand(wall_particles.values, instance(wall_particles))
    
    particle_velocity= math.concat([fluid_particle_velocity,wall_particle_velocity], 'particles')
 
    vmax_magnitude= math.max(math.vec_length(particle_velocity))

    c_max= fluid_c_0  # wall_c_0 =0 so no point in taking the max out of them

    dt_1=0.25*h/(c_max+vmax_magnitude) # single value

    #viscous condition
    mu=0.5/(d+2) *fluid_alpha*h*c_max
    dt_2=0.125*(h**2)/mu

    dt_3= 0.25*math.sqrt(h/abs(g))

    dt= tensor([dt_1,dt_2,dt_3], instance('time_steps') )

    dt = math.min(dt)

    return dt

def kernal_function(d,h,q):
    # Quintic Spline used as Weighting function
    # cutoff radius r_c = 3 * h;
    alpha_d = 0.004661441847880 / (h * h)

    # Weighting function
    W= math.where((q<3) & (q>=2) ,  alpha_d * ((3-q)**5)                                     ,q)
    W= math.where((q<2) & (q>=1) ,  alpha_d * (((3-q)**5) - 6 * ((2-q)**5))                  ,W)
    W= math.where((q<1) & (q>0)  ,  alpha_d * (((3-q)**5) - 6 * ((2-q)**5) + 15 * ((1-q)**5)),W)
    W= math.where((q>=3)         ,  0                                                        ,W)

    return W

def kernal_derivative(d, h, q):
    alpha_d = -0.0233072092393989 / (h * h)

    der_W= math.where((q<3) & (q>=2) ,  alpha_d * ((3-q)**4)                                     ,q)
    der_W= math.where((q<2) & (q>=1) ,  alpha_d * (((3-q)**4) - 6 * ((2-q)**4))                  ,der_W)
    der_W= math.where((q<1) & (q>0)  ,  alpha_d * (((3-q)**4) - 6 * ((2-q)**4) + 15 * ((1-q)**4)),der_W)
    der_W= math.where((q>=3)         ,  0                                                        ,der_W)
    
    return der_W

def intial_condition(dx, d, alph):
    width = dx * np.ceil(1.61 / dx)  # width=1.614 for dx= 0.006
    g = -9.81
    ###CHANGE LATER TO 0.3
    height = 0.3
    v_max = np.sqrt(2 * abs(g) * height)
    
    ###### Properties of Fluid Particles #######
    fluid_initial_density = 1000.0
    fluid_adiabatic_exp = 7.0  # adiabatic coefficient (pressure coefficient)
    fluid_c_0=10.0*v_max # artificial speed of sound c_0 and v_max = 2*abs(g)*height
    fluid_p_0=(fluid_initial_density*((fluid_c_0)**2))/fluid_adiabatic_exp  # reference pressure 
    fluid_Xi=0.0  # background pressure
    fluid_mu=0.01  # viscosity
    fluid_alpha=alph  # artificial visc factor

    #####UNCOMMENT LATER
    fluid_coords = pack_dims(math.meshgrid(x=100, y=50), 'x,y', instance('particles')) * (0.6/100, 0.3/50) + (0.003,0.003)  # 5000 fluid particle coordinates created     

    #fluid_coords = pack_dims(math.meshgrid(x=25, y=25), 'x,y', instance('particles')) * (0.15/25, 0.15/25) + (0.825,0.005)  # 625 fluid particle coordinates created     
    # fluid_coords_1 =( pack_dims(math.meshgrid(x=25, y=25), 'x,y', instance('particles')) * (0.15/25, 0.15/25) + (1.467,0.005))
    # fluid_coords_2 = ( pack_dims(math.meshgrid(x=25, y=25), 'x,y', instance('particles')) * (0.15/25, 0.15/25) + (0.003,0.005))
    # fluid_coords= math.concat([fluid_coords_1, fluid_coords_2 ], 'particles')


    #fluid_coords = pack_dims(math.meshgrid(x=3, y=2), 'x,y', instance('particles')) * (0.018/3.0, 0.012/2.0) + (0.825,0.10)  # 9 fluid particle coordinates created     
    #fluid_coords = pack_dims(math.meshgrid(x=1, y=1), 'x,y', instance('particles')) * (0.006/1.0, 0.006/1.0) + (0.825,0.050)  # 1 fluid particle coordinates created     
    #fluid_coords = pack_dims(math.meshgrid(x=25, y=1), 'x,y', instance('particles')) * (0.15/25, 1) + (0.003,0.003)  # 5000 fluid particle coordinates created     
    #fluid_coords =pack_dims(math.meshgrid(x=1, y=3), 'x,y', instance('particles')) * (0.2/1, 0.12/3) + (0.003,0.003)
    
    fluid_particles = PointCloud(Sphere(fluid_coords, radius=0.01))  #"""is this radius only for visualization?????????????"""
    #math.print(math.zeros(instance(fluid_particles.elements.center)))
    
    fluid_velocity = fluid_particles * (0.0, 0.0)  # can we remove this unnecessary point cloud creation ?
    fluid_particles = fluid_particles.with_values(fluid_velocity.values)  #fluid particles is a point cloud with elements as points of fluid coordinates and values as velocity

    single_fluid_particle_mass = fluid_initial_density * dx**d
    fluid_particle_mass = math.ones(instance(fluid_coords))*single_fluid_particle_mass
    
    fluid_pressure=math.zeros(instance(fluid_coords))  

    ###### Properties of Wall particles #####
    wall_initial_density =0.0 
    wall_adiabatic_exp = 0.0  # adiabatic coefficient (pressure coefficient)
    wall_c_0=0.0 # artificial speed of sound c_0
    wall_p_0=0.0  # reference pressure 
    wall_Xi=0.0  # background pressure
    wall_mu=0.0  # viscosity
    wall_alpha=0-0  # artificial visc factor

    left_wall_coords = pack_dims(math.meshgrid(x=3, y=134), 'x,y', instance('particles')) * ( (0.018/3), (0.804/134) ) + (-0.015, 0.003)  
    # #print(f"{left_wall_coords:full:shape}")
    right_wall_coords = pack_dims(math.meshgrid(x=3, y=134), 'x,y', instance('particles')) * ( (0.018/3), (0.804/134) ) + (1.617, 0.003)  

    center_wall_coords = (pack_dims(math.meshgrid(x=275, y=3), 'x,y', instance('particles')) * ( (1.65/275), (0.018/3) ) + (-0.015, -0.015))
    
    #center_wall_coords = (pack_dims(math.meshgrid(x=5, y=3), 'x,y', instance('particles')) * ( (0.6/100), (0.018/3) ) + (0.815,-0.015))
    #center_wall_coords = (pack_dims(math.meshgrid(x=5, y=3), 'x,y', instance('particles')) * ( (0.6/100), (0.018/3) ) + (-0.015, -0.015))


    # concatenating the wall coordinates
    wall_coords=math.concat([left_wall_coords, right_wall_coords,center_wall_coords], 'particles') #1629 wall particles
    
    wall_particles = PointCloud(Sphere(wall_coords, radius=0.01), color='#FFA500')

    wall_initial_velocity=wall_particles * (0, 0)
    wall_particles = wall_particles.with_values(wall_initial_velocity.values) 
    wall_pressure=math.zeros(instance(wall_coords)) 
    wall_density=math.zeros(instance(wall_coords))  

    #particles[:number_fluid_particles,8]=rho_0[0]*abs(g)*(height-particles[:number_fluid_particles,1])
    fluid_pressure = fluid_initial_density * abs(g) * (height - fluid_particles.points['y']) 

    #particles[:number_fluid_particles, 5] = rho_0[0] * (((particles[:number_fluid_particles, 8] - Xi[0]) / p_0[0]) + 1) ** (1 / gamma[0])
    fluid_density = fluid_initial_density * (((fluid_pressure-fluid_Xi)/fluid_p_0)+1)**(1.0/fluid_adiabatic_exp)
    

    return fluid_particles,wall_particles, fluid_initial_density,wall_initial_density, \
    fluid_density, wall_density,fluid_pressure, wall_pressure,  \
    fluid_particle_mass, \
    fluid_adiabatic_exp,wall_adiabatic_exp, fluid_c_0,wall_c_0,fluid_p_0,wall_p_0,fluid_Xi,wall_Xi,fluid_alpha,wall_alpha, \
    g, height




if __name__== "__main__":
    main()