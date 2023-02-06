import pandas as pd
from openpyxl import Workbook
from phi.flow import *
from sph_phiflow import *


def main():






        t = t + dt
        n_dt = n_dt + 1

        if n_dt == 1:


            # workbook_y = Workbook()
            # workbook_y.save("y_pos.xlsx")
            #
            # workbook_x = Workbook()
            # workbook_x.save("x_pos.xlsx")
            #
            # pos_x = numpy.asarray(math.concat([fluid_particles.points['x'], wall_particles.points['x']], 'particles'))
            # pos_y = numpy.asarray(math.concat([fluid_particles.points['y'], wall_particles.points['y']], 'particles'))
            # df_x = pd.DataFrame(pos_x)
            # df_y = pd.DataFrame(pos_y)

            # with pd.ExcelWriter("x_pos.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer_x:
            #     df_x.to_excel(writer_x, sheet_name="Sheet", header=None, startcol=writer_x.sheets["Sheet"].max_column, index=False)
            #     writer_x.save()
            #
            # with pd.ExcelWriter("y_pos.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer_y:
            #     df_y.to_excel(writer_y, sheet_name="Sheet", header=None, startcol=writer_y.sheets["Sheet"].max_column, index=False)
            #     writer_y.save()

        # *******************************************************************************************
        # Step-Function: Calculate the parameters for one step
        # *******************************************************************************************

        # *******************************************************************************************

        # fluid_traj.append(fluid_particles)

        # if (n_dt % 150 == 0):
        #     pos_x = numpy.asarray(math.concat([fluid_particles.points['x'], wall_particles.points['x']], 'particles'))
        #     pos_y = numpy.asarray(math.concat([fluid_particles.points['y'], wall_particles.points['y']], 'particles'))
        #     df_x = pd.DataFrame(pos_x)
        #     df_y = pd.DataFrame(pos_y)
        #
        #     with pd.ExcelWriter("x_pos.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer_x:
        #         df_x.to_excel(writer_x, sheet_name="Sheet", header=None, startcol=writer_x.sheets["Sheet"].max_column, index=False)
        #         writer_x.save()
        #
        #     with pd.ExcelWriter("y_pos.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer_y:
        #         df_y.to_excel(writer_y, sheet_name="Sheet", header=None, startcol=writer_y.sheets["Sheet"].max_column, index=False)
        #         writer_y.save()

    # **************************
    # Save animation
    # **************************
    # fluid_trj = math.stack(fluid_traj, batch('time'))
    # a: matplotlib.animation.FuncAnimation = vis.plot(vis.overlay(wall_particles.elements, fluid_trj.elements), animate='time')
    # a.save('anim.mp4')


def free_fall_case(dx, d, alph):
    width = dx * np.ceil(1.61 / dx)  # width=1.614 for dx= 0.006
    g = -9.81
    ###CHANGE LATER TO 0.3
    height = 0.15
    v_max = np.sqrt(2 * abs(g) * height)

    ###### Properties of Fluid Particles #######
    fluid_initial_density = 1000.0
    fluid_adiabatic_exp = 7.0  # adiabatic coefficient (pressure coefficient)
    fluid_c_0 = 10.0 * v_max  # artificial speed of sound c_0 and v_max = 2*abs(g)*height
    fluid_p_0 = (fluid_initial_density * ((fluid_c_0) ** 2)) / fluid_adiabatic_exp  # reference pressure
    fluid_Xi = 0.0  # background pressure
    fluid_mu = 0.01  # viscosity
    fluid_alpha = alph  # artificial visc factor

    #####UNCOMMENT LATER
    # fluid_coords = pack_dims(math.meshgrid(x=100, y=50), 'x,y', instance('particles')) * (0.6/100, 0.3/50) + (0.003,0.003)  # 5000 fluid particle coordinates created

    fluid_coords = pack_dims(math.meshgrid(x=25, y=25), 'x,y', instance('particles')) * (0.15 / 25, 0.15 / 25) + (0.825, 0.005)  # 625 fluid particle coordinates created
    # fluid_coords = pack_dims(math.meshgrid(x=3, y=2), 'x,y', instance('particles')) * (0.018/3.0, 0.012/2.0) + (0.825,0.10)  # 9 fluid particle coordinates created
    # fluid_coords = pack_dims(math.meshgrid(x=1, y=1), 'x,y', instance('particles')) * (0.006/1.0, 0.006/1.0) + (0.825,0.050)  # 1 fluid particle coordinates created
    # fluid_coords = pack_dims(math.meshgrid(x=25, y=1), 'x,y', instance('particles')) * (0.15/25, 1) + (0.003,0.003)  # 5000 fluid particle coordinates created
    # fluid_coords =pack_dims(math.meshgrid(x=1, y=3), 'x,y', instance('particles')) * (0.2/1, 0.12/3) + (0.003,0.003)

    fluid_particles = PointCloud(Sphere(fluid_coords, radius=0.1))  # """is this radius only for visualization?????????????"""
    # math.print(math.zeros(instance(fluid_particles.elements.center)))

    fluid_velocity = fluid_particles * (0.0, 0.0)  # can we remove this unnecessary point cloud creation ?
    fluid_particles = fluid_particles.with_values(fluid_velocity.values)  # fluid particles is a point cloud with elements as points of fluid coordinates and values as velocity

    single_fluid_particle_mass = fluid_initial_density * dx ** d
    fluid_particle_mass = math.ones(instance(fluid_coords)) * single_fluid_particle_mass

    fluid_pressure = math.zeros(instance(fluid_coords))

    ###### Properties of Wall particles #####
    wall_initial_density = 0.0
    wall_adiabatic_exp = 0.0  # adiabatic coefficient (pressure coefficient)
    wall_c_0 = 0.0  # artificial speed of sound c_0
    wall_p_0 = 0.0  # reference pressure
    wall_Xi = 0.0  # background pressure
    wall_mu = 0.0  # viscosity
    wall_alpha = 0 - 0  # artificial visc factor

    left_wall_coords = pack_dims(math.meshgrid(x=3, y=134), 'x,y', instance('particles')) * ((0.018 / 3), (0.804 / 134)) + (-0.015, 0.003)
    # #print(f"{left_wall_coords:full:shape}")
    right_wall_coords = pack_dims(math.meshgrid(x=3, y=134), 'x,y', instance('particles')) * ((0.018 / 3), (0.804 / 134)) + (1.617, 0.003)

    center_wall_coords = (pack_dims(math.meshgrid(x=275, y=3), 'x,y', instance('particles')) * ((0.6 / 100), (0.018 / 3)) + (-0.015, -0.015))

    # center_wall_coords = (pack_dims(math.meshgrid(x=5, y=3), 'x,y', instance('particles')) * ( (0.6/100), (0.018/3) ) + (0.815,-0.015))
    # center_wall_coords = (pack_dims(math.meshgrid(x=5, y=3), 'x,y', instance('particles')) * ( (0.6/100), (0.018/3) ) + (-0.015, -0.015))

    # concatenating the wall coordinates
    wall_coords = math.concat([left_wall_coords, right_wall_coords, center_wall_coords], 'particles')  # 1629 wall particles

    wall_particles = PointCloud(Sphere(wall_coords, radius=0.002))

    wall_initial_velocity = wall_particles * (0, 0)
    wall_particles = wall_particles.with_values(wall_initial_velocity.values)
    wall_pressure = math.zeros(instance(wall_coords))
    wall_density = math.zeros(instance(wall_coords))

    # particles[:number_fluid_particles,8]=rho_0[0]*abs(g)*(height-particles[:number_fluid_particles,1])
    fluid_pressure = fluid_initial_density * abs(g) * (height - fluid_particles.points['y'])

    # particles[:number_fluid_particles, 5] = rho_0[0] * (((particles[:number_fluid_particles, 8] - Xi[0]) / p_0[0]) + 1) ** (1 / gamma[0])
    fluid_density = fluid_initial_density * (((fluid_pressure - fluid_Xi) / fluid_p_0) + 1) ** (1.0 / fluid_adiabatic_exp)

    return fluid_particles, wall_particles, fluid_initial_density, wall_initial_density, \
           fluid_density, wall_density, fluid_pressure, wall_pressure, \
           fluid_particle_mass, \
           fluid_adiabatic_exp, wall_adiabatic_exp, fluid_c_0, wall_c_0, fluid_p_0, wall_p_0, fluid_Xi, wall_Xi, fluid_alpha, wall_alpha, \
           g, height


if __name__ == "__main__":
    main()
